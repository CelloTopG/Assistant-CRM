import json
from datetime import date
from typing import Any, Dict, List, Tuple, Optional

import frappe
from frappe.model.document import Document
from frappe.utils import getdate, now_datetime, add_days


class BeneficiaryStatusReport(Document):
    def before_insert(self):
        self._ensure_dates()

    def before_save(self):
        # Keep dates sane but don't auto-run heavy aggregation here
        self._ensure_dates()

    def _is_monthly_window(self) -> bool:
        try:
            if getattr(self, "period_type", None) != "Monthly":
                return False
            df = frappe.utils.getdate(self.date_from)
            dt = frappe.utils.getdate(self.date_to)
            return df == frappe.utils.get_first_day(df) and dt == frappe.utils.get_last_day(df)
        except Exception:
            return False

    def _month_bounds(self):
        df = frappe.utils.getdate(self.date_from)
        return frappe.utils.get_first_day(df), frappe.utils.get_last_day(df)

    @frappe.whitelist()
    def run_generation(self):
        self._ensure_dates()
        try:
            if self._is_monthly_window():
                counts, rows, charts = get_or_build_monthly_snapshot(self.date_from, self.date_to)
            else:
                counts, rows, charts = aggregate_beneficiaries_sql(self.date_from, self.date_to)
        except Exception:
            # Fallback to legacy in-memory aggregation
            counts, rows, charts = aggregate_beneficiaries(self.date_from, self.date_to)
        self.total_beneficiaries = counts.get("total", 0)
        self.active_count = counts.get("active", 0)
        self.suspended_count = counts.get("suspended", 0)
        self.deceased_count = counts.get("deceased", 0)
        self.pending_verification_count = counts.get("pending_verification", 0)
        self.terminated_count = counts.get("terminated", 0)
        self.status_chart_json = json.dumps(charts.get("status") or {})
        self.province_chart_json = json.dumps(charts.get("province") or {})
        self.benefit_type_chart_json = json.dumps(charts.get("benefit_type") or {})
        self.trend_chart_json = json.dumps(charts.get("trend") or {})
        self.rows_json = json.dumps(rows[:500])
        self.report_html = build_report_html(self, counts)
        self.generated_at = now_datetime()
        self.generated_by = frappe.session.user
        self.save()
        return {"ok": True, "counts": counts}

    @frappe.whitelist()
    def generate_pdf(self) -> str:
        from frappe.utils.pdf import get_pdf
        html = _build_pdf_html(self)
        filename = f"Beneficiary_Status_Report_{self.name}.pdf"
        _attach_to_doc(self.doctype, self.name, filename, get_pdf(html))
        return filename

    @frappe.whitelist()
    def generate_excel(self) -> str:
        rows = json.loads(self.rows_json or "[]")
        filename = f"Beneficiary_Status_Report_{self.name}.xlsx"
        content: bytes
        # Prepare status partitions
        statuses = ["Active", "Suspended", "Deceased", "Pending Verification", "Terminated"]
        headers = ["Beneficiary #", "Full Name", "Benefit Status", "Life Status", "Province", "Benefit Type"]
        by_status: Dict[str, List[Dict[str, Any]]] = {s: [] for s in statuses}
        for r in rows:
            s = (r.get("benefit_status") or "").strip()
            if s in by_status:
                by_status[s].append(r)
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            # Title
            ws["A1"] = "WCFCB Beneficiary Status Report"
            ws["A1"].font = Font(b=True, size=14)
            ws["A2"] = f"{self.period_type} | {self.date_from} → {self.date_to}"
            ws["A2"].font = Font(color="666666")

            # Counts table
            ws["A4"] = "Metric"; ws["B4"] = "Count"
            ws["A4"].font = ws["B4"].font = Font(b=True)
            metrics = [
                ("Total", int(self.total_beneficiaries or 0)),
                ("Active", int(self.active_count or 0)),
                ("Suspended", int(self.suspended_count or 0)),
                ("Deceased", int(self.deceased_count or 0)),
                ("Pending Verification", int(self.pending_verification_count or 0)),
                ("Terminated", int(self.terminated_count or 0)),
            ]
            r = 5
            for k, v in metrics:
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=v)
                r += 1

            # Legend
            ws["D4"] = "Legend"; ws["D4"].font = Font(b=True)
            legend = [
                ("Active", "E9F7EF"), ("Suspended", "FEF9E7"), ("Deceased", "FDEDEC"), ("Pending Verification", "EBF5FB"), ("Terminated", "FDECEA"),
            ]
            rr = 5
            for name, color in legend:
                ws.cell(row=rr, column=4, value=name)
                c = ws.cell(row=rr, column=5, value=" ")
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                rr += 1

            # Per-status sheets
            fills = {
                "Active": PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid"),
                "Suspended": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
                "Deceased": PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
                "Pending Verification": PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid"),
                "Terminated": PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid"),
            }

            def write_sheet(title: str, data: List[Dict[str, Any]]):
                wsx = wb.create_sheet(title=title[:31])
                for ci, h in enumerate(headers, start=1):
                    cell = wsx.cell(row=1, column=ci, value=h)
                    cell.font = Font(b=True)
                rr = 2
                for row in data:
                    vals = [
                        row.get("beneficiary_number"), row.get("full_name"), row.get("benefit_status"),
                        row.get("life_status"), row.get("province"), row.get("benefit_type")
                    ]
                    for ci, v in enumerate(vals, start=1):
                        wsx.cell(row=rr, column=ci, value=v)
                    # Row fill based on status
                    for ci in range(1, len(headers) + 1):
                        wsx.cell(row=rr, column=ci).fill = fills.get(title, PatternFill())
                    rr += 1

            for s in statuses:
                write_sheet(s, by_status.get(s) or [])

            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()
        except Exception:
            # Fallback to single-sheet via xlsxutils; last fallback CSV
            rows_simple = [
                [r.get("beneficiary_number"), r.get("full_name"), r.get("benefit_status"), r.get("life_status"), r.get("province"), r.get("benefit_type")] for r in rows
            ]
            try:
                from frappe.utils.xlsxutils import make_xlsx  # type: ignore
                xlsx_file = make_xlsx([[*headers]] + rows_simple, "Beneficiaries")
                content = xlsx_file.getvalue()
            except Exception:
                import io, csv
                buf = io.StringIO()
                w = csv.writer(buf)
                w.writerow(headers)
                w.writerows(rows_simple)
                content = buf.getvalue().encode()
                filename = filename.replace(".xlsx", ".csv")
        _attach_to_doc(self.doctype, self.name, filename, content)
        return filename

    @frappe.whitelist()
    def email_report(self, recipients: Optional[List[str]] = None) -> Dict[str, Any]:
        recips = recipients or _get_role_emails(["Accounts Manager", "Finance Manager"]) or []
        if not recips:
            return {"message": "no-recipients"}
        try:
            # Ensure latest aggregation
            self.run_generation()
            # Attach PDF and Excel
            pdf_name = self.generate_pdf()
            xlsx_name = self.generate_excel()
            files = frappe.get_all("File", filters={"attached_to_doctype": self.doctype, "attached_to_name": self.name}, fields=["name", "file_name", "content", "file_url"], limit=10)
            # Fetch file contents
            attachments = []
            for f in files:
                file_doc = frappe.get_doc("File", f.get("name"))
                attachments.append({"fname": file_doc.file_name, "fcontent": file_doc.get_content()})
            frappe.sendmail(
                recipients=recips,
                subject=f"WCFCB Beneficiary Status Report: {self.period_type} ({self.date_from} → {self.date_to})",
                message="Attached is the latest Beneficiary Status Report.",
                attachments=attachments,
            )
            return {"message": "sent", "recipients": recips}
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Beneficiary Status Report: email_report")
            frappe.throw(f"Failed to email report: {e}")


# ----- Aggregation -----

def aggregate_beneficiaries(date_from: Any, date_to: Any) -> Tuple[Dict[str, int], List[Dict[str, Any]], Dict[str, Any]]:
    df = getdate(date_from)
    dt = getdate(date_to)

    # Query the native Beneficiary Profile doctype directly
    fields = [
        "name", "beneficiary_number", "first_name", "last_name",
        "benefit_status", "life_status", "province", "benefit_type",
        "benefit_start_date", "benefit_end_date"
    ]
    rows = frappe.get_all("Beneficiary Profile", fields=fields, limit=20000)

    counts = {"total": 0, "active": 0, "suspended": 0, "deceased": 0, "pending_verification": 0, "terminated": 0}
    status_rows: List[Dict[str, Any]] = []
    province_map: Dict[str, int] = {}
    benefit_type_map: Dict[str, int] = {}

    for r in rows:
        counts["total"] += 1
        bs = (r.get("benefit_status") or "").strip()
        ls = (r.get("life_status") or "").strip()
        # Active window: started by dt and not ended before df
        started = not r.get("benefit_start_date") or getdate(r["benefit_start_date"]) <= dt
        not_ended = not r.get("benefit_end_date") or getdate(r["benefit_end_date"]) >= df
        is_active_window = started and not_ended

        if ls == "Deceased" or bs == "Deceased":
            counts["deceased"] += 1
        elif bs == "Suspended":
            counts["suspended"] += 1
        elif bs in {"Pending", "Under Review", "Pending Verification"}:
            counts["pending_verification"] += 1
        elif bs == "Terminated":
            counts["terminated"] += 1
        elif bs == "Active" and is_active_window:
            counts["active"] += 1

        province = r.get("province") or "Unknown"
        province_map[province] = province_map.get(province, 0) + 1
        btype = r.get("benefit_type") or "Unknown"
        benefit_type_map[btype] = benefit_type_map.get(btype, 0) + 1

        status_rows.append({
            "beneficiary_number": r.get("beneficiary_number"),
            "full_name": " ".join([x for x in [r.get("first_name"), r.get("last_name")] if x]),
            "benefit_status": bs,
            "life_status": ls or "",
            "province": province,
            "benefit_type": btype,
        })

    charts = {
        "status": {
            "data": {
                "labels": ["Active", "Suspended", "Deceased", "Pending Verification", "Terminated"],
                "datasets": [{
                    "name": "Beneficiaries",
                    "values": [counts.get("active", 0), counts.get("suspended", 0), counts.get("deceased", 0), counts.get("pending_verification", 0), counts.get("terminated", 0)]
                }]
            },
            "type": "pie"
        },
        "province": {
            "data": {
                "labels": list(province_map.keys()),
                "datasets": [{"name": "Count", "chartType": "bar", "values": [province_map[k] for k in province_map.keys()]}]
            },
            "type": "bar"
        },
        "benefit_type": {
            "data": {
                "labels": list(benefit_type_map.keys()),
                "datasets": [{"name": "Count", "chartType": "bar", "values": [benefit_type_map[k] for k in benefit_type_map.keys()]}]
            },
            "type": "bar"
        },
        # Simple placeholder trend: last 6 months total (re-aggregated if needed later)
        "trend": {
            "data": {"labels": [], "datasets": [{"name": "Total", "values": []}]},
            "type": "line"
        }
    }


# ----- Optimized Aggregation + Snapshots -----

def _first_day(d):
    return frappe.utils.get_first_day(frappe.utils.getdate(d))


def _last_day(d):
    return frappe.utils.get_last_day(frappe.utils.getdate(d))


def get_or_build_monthly_snapshot(date_from: Any, date_to: Any):
    """Build monthly snapshot data - snapshot caching has been deprecated"""
    dt = _last_day(date_to)
    # Build fresh data (snapshot caching removed)
    counts, rows, charts, maps = aggregate_beneficiaries_sql(date_from, date_to, return_maps=True)
    return counts, rows, charts


def aggregate_beneficiaries_sql(date_from: Any, date_to: Any, return_maps: bool = False):
    """Aggregate beneficiaries using SQL for performance.

    Uses the native Beneficiary Profile doctype directly.
    """
    df = getdate(date_from)
    dt = getdate(date_to)
    t = "`tabBeneficiary Profile`"

    # Counts using indexed filters
    def q(sql, params=None):
        return frappe.db.sql(sql, params or {}, as_dict=False)[0][0]

    counts = {}
    counts["total"] = q(f"SELECT COUNT(*) FROM {t}")
    counts["active"] = q(
        f"""
        SELECT COUNT(*) FROM {t}
        WHERE benefit_status='Active'
          AND (benefit_start_date IS NULL OR benefit_start_date <= %(dt)s)
          AND (benefit_end_date IS NULL OR benefit_end_date >= %(df)s)
        """,
        {"df": df, "dt": dt},
    )
    counts["suspended"] = q(f"SELECT COUNT(*) FROM {t} WHERE benefit_status='Suspended'")
    counts["deceased"] = q(
        f"SELECT COUNT(*) FROM {t} WHERE life_status='Deceased' OR benefit_status='Deceased'"
    )
    counts["pending_verification"] = q(
        f"SELECT COUNT(*) FROM {t} WHERE benefit_status IN ('Pending Verification','Under Review','Pending')"
    )
    counts["terminated"] = q(f"SELECT COUNT(*) FROM {t} WHERE benefit_status='Terminated'")

    # Distributions
    def qmap(sql, key):
        out = {}
        for k, v in frappe.db.sql(sql, as_dict=False):
            out[k or "Unknown"] = int(v)
        return out
    province_map = qmap(f"SELECT IFNULL(province,'Unknown') AS p, COUNT(*) FROM {t} GROUP BY p", "province")
    benefit_type_map = qmap(f"SELECT IFNULL(benefit_type,'Unknown') AS b, COUNT(*) FROM {t} GROUP BY b", "benefit_type")

    # Sample rows for UI/Excel tabs
    rows_raw = frappe.db.sql(
        f"""
        SELECT beneficiary_number, first_name, last_name, benefit_status, life_status, province, benefit_type
        FROM {t}
        ORDER BY modified DESC
        LIMIT 2000
        """,
        as_dict=True,
    )
    status_rows = [
        {
            "beneficiary_number": r.get("beneficiary_number"),
            "full_name": " ".join([x for x in [r.get("first_name"), r.get("last_name")] if x]),
            "benefit_status": r.get("benefit_status") or "",
            "life_status": r.get("life_status") or "",
            "province": r.get("province") or "Unknown",
            "benefit_type": r.get("benefit_type") or "Unknown",
        }
        for r in rows_raw
    ]

    charts = _build_charts_from_maps(counts, province_map, benefit_type_map, dt)
    if return_maps:
        return counts, status_rows, charts, {"province": province_map, "benefit_type": benefit_type_map}
    return counts, status_rows, charts


def _build_charts_from_maps(counts: Dict[str, int], province_map: Dict[str, int], benefit_type_map: Dict[str, int], anchor_date):
    charts = {
        "status": {
            "data": {
                "labels": ["Active", "Suspended", "Deceased", "Pending Verification", "Terminated"],
                "datasets": [{
                    "name": "Beneficiaries",
                    "values": [
                        counts.get("active", 0), counts.get("suspended", 0), counts.get("deceased", 0),
                        counts.get("pending_verification", 0), counts.get("terminated", 0)
                    ]
                }]
            },
            "type": "pie",
        },
        "province": {
            "data": {
                "labels": list(province_map.keys()),
                "datasets": [{"name": "Count", "chartType": "bar", "values": [province_map[k] for k in province_map.keys()]}]
            },
            "type": "bar",
        },
        "benefit_type": {
            "data": {
                "labels": list(benefit_type_map.keys()),
                "datasets": [{"name": "Count", "chartType": "bar", "values": [benefit_type_map[k] for k in benefit_type_map.keys()]}]
            },
            "type": "bar",
        },
        "trend": {"data": {"labels": [], "datasets": [{"name": "Total", "values": []}]}, "type": "line"},
    }
    # Build 6-month trend using current total from Beneficiary Profile
    labels: List[str] = []
    values: List[int] = []
    anchor = getdate(anchor_date)
    total = frappe.db.count("Beneficiary Profile")
    for i in range(5, -1, -1):
        mstart = frappe.utils.add_months(frappe.utils.get_first_day(anchor), -i)
        labels.append(mstart.strftime("%b %Y"))
        values.append(int(total))
    charts["trend"]["data"]["labels"] = labels
    charts["trend"]["data"]["datasets"][0]["values"] = values
    return charts



# ----- AI Insights -----
@frappe.whitelist()
def get_ai_insights(name: str, query: str) -> Dict[str, Any]:
    """Return WorkCom-style insights for a Beneficiary Status Report.

    Builds a JSON context with status counts, distributions and recent
    historical snapshots, and passes it to WorkCom via EnhancedAIService.
    """
    doc = frappe.get_doc("Beneficiary Status Report", name)

    history = frappe.get_all(
        "Beneficiary Status Report",
        filters={"period_type": doc.period_type},
        fields=[
            "name",
            "date_from",
            "date_to",
            "total_beneficiaries",
            "active_count",
            "suspended_count",
            "deceased_count",
            "pending_verification_count",
            "terminated_count",
        ],
        order_by="date_from desc",
        limit=10,
    )

    # Reuse same aggregation logic as run_generation so WorkCom sees
    # the latest numbers and distributions
    try:
        if doc._is_monthly_window():
            counts, rows, charts = get_or_build_monthly_snapshot(doc.date_from, doc.date_to)
        else:
            counts, rows, charts = aggregate_beneficiaries_sql(doc.date_from, doc.date_to)
    except Exception:
        counts, rows, charts = aggregate_beneficiaries(doc.date_from, doc.date_to)

    context = {
        "window": {
            "period_type": doc.period_type,
            "from": str(doc.date_from),
            "to": str(doc.date_to),
        },
        "current": {
            "total": counts.get("total", int(doc.total_beneficiaries or 0)),
            "active": counts.get("active", int(doc.active_count or 0)),
            "suspended": counts.get("suspended", int(doc.suspended_count or 0)),
            "deceased": counts.get("deceased", int(doc.deceased_count or 0)),
            "pending_verification": counts.get("pending_verification", int(doc.pending_verification_count or 0)),
            "terminated": counts.get("terminated", int(doc.terminated_count or 0)),
        },
        "distributions": {
            "status_chart": json.loads(doc.status_chart_json or "{}"),
            "province_chart": json.loads(doc.province_chart_json or "{}"),
            "benefit_type_chart": json.loads(doc.benefit_type_chart_json or "{}"),
        },
        "history": history,
    }

    try:
        from assistant_crm.services.enhanced_ai_service import EnhancedAIService

        ai = EnhancedAIService()
        answer = ai.generate_beneficiary_status_report_insights(query=query, context=context)
        return {"insights": answer}
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Beneficiary Status Report AI Insights Error")
        return {
            "insights": (
                "AI insights are temporarily unavailable. Please ask your system "
                "administrator to configure WorkCom/OpenAI settings in Enhanced AI Settings."
            )
        }


# ----- Scheduler -----

def _is_first_business_day(d: date) -> bool:
    wd = d.weekday()
    first = frappe.utils.get_first_day(d)
    first_bd = first
    while first_bd.weekday() > 4:
        first_bd = frappe.utils.add_days(first_bd, 1)
    return d == first_bd


def _previous_month(d: date) -> Tuple[str, int]:
    prev = frappe.utils.add_months(d, -1)
    return prev.strftime("%b"), prev.year


@frappe.whitelist()
def schedule_monthly_beneficiary_status_reports():
    today = frappe.utils.getdate()
    if _is_first_business_day(today):
        # create previous month doc and email
        mon, yr = _previous_month(today)
        df = frappe.utils.get_first_day(f"{yr}-{mon}-01")
        dt = frappe.utils.get_last_day(df)
        doc = frappe.new_doc("Beneficiary Status Report")
        doc.period_type = "Monthly"
        doc.date_from = df
        doc.date_to = dt
        doc.title = f"Beneficiaries Monthly: {df} → {dt}"
        doc.insert(ignore_permissions=True)
        doc.run_generation()
        # send to Finance roles
        doc.email_report()


# ----- Helpers -----

def _attach_to_doc(doctype: str, name: str, filename: str, content: bytes):
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "attached_to_doctype": doctype,
        "attached_to_name": name,
        "content": content,
        "is_private": 1,
    })
    file_doc.save(ignore_permissions=True)


def _get_role_emails(roles: List[str]) -> List[str]:
    if not roles:
        return []
    has_roles = frappe.get_all("Has Role", filters={"role": ["in", roles]}, fields=["parent"], limit=1000)
    user_ids = {r["parent"] for r in has_roles if r.get("parent")}
    if not user_ids:
        return []
    users = frappe.get_all("User", filters={"name": ["in", list(user_ids)], "enabled": 1, "user_type": "System User"}, fields=["email"], limit=1000)
    return sorted({u.get("email") for u in users if u.get("email")})


def build_report_html(doc: "BeneficiaryStatusReport", counts: Dict[str, int]) -> str:
    def card(label: str, value: Any, style: str = ""):
        return f"<div style='display:inline-block;margin:6px;padding:10px;border:1px solid #ddd;border-radius:6px{style}'><div style='font-size:12px;color:#666'>{label}</div><div style='font-size:18px;font-weight:600'>{value}</div></div>"

    cards = "".join([
        card("Total", counts.get("total", 0)),
        card("Active", counts.get("active", 0)),
        card("Suspended", counts.get("suspended", 0)),
        card("Deceased", counts.get("deceased", 0)),
        card("Pending Verification", counts.get("pending_verification", 0)),
        card("Terminated", counts.get("terminated", 0), style=":background:#ffe5e5"),
    ])
    return f"<div><div style='margin-bottom:10px'>{cards}</div></div>"


def _build_pdf_html(doc: "BeneficiaryStatusReport") -> str:
    subtitle = f"{doc.period_type} | {doc.date_from} → {doc.date_to}"
    header = f"<h2 style='margin:0'>WCFCB Beneficiary Status Report</h2><div style='color:#666'>{subtitle}</div>"
    body = build_report_html(doc, {
        "total": doc.total_beneficiaries or 0,
        "active": doc.active_count or 0,
        "suspended": doc.suspended_count or 0,
        "deceased": doc.deceased_count or 0,
        "pending_verification": doc.pending_verification_count or 0,
        "terminated": doc.terminated_count or 0,
    })
    return f"<div style='font-family:Inter,Arial,sans-serif'>{header}<hr/>{body}</div>"


# ----- Date Helpers Bound to Class -----

def _ensure_dates(self: Document):
    if not getattr(self, "date_from", None) or not getattr(self, "date_to", None):
        if getattr(self, "period_type", None) == "Monthly":
            start = frappe.utils.get_first_day(frappe.utils.getdate())
            end = frappe.utils.getdate()
        else:
            # Custom: default to last 30 days
            end = frappe.utils.getdate()
            start = add_days(end, -29)
        self.date_from, self.date_to = start, end


BeneficiaryStatusReport._ensure_dates = _ensure_dates




@frappe.whitelist()
def smoke_generate_report(period_type: str = "Monthly", date_from: Optional[str] = None, date_to: Optional[str] = None) -> str:
    """Convenience helper to create, generate and attach files for a report. Returns docname.
    If Monthly and no dates provided, uses previous full month.
    """
    if not date_from or not date_to:
        today = frappe.utils.getdate()
        # default to previous month window
        prev = frappe.utils.add_months(frappe.utils.get_first_day(today), -1)
        df = frappe.utils.get_first_day(prev)
        dt = frappe.utils.get_last_day(prev)
    else:
        df = frappe.utils.getdate(date_from)
        dt = frappe.utils.getdate(date_to)

    doc = frappe.new_doc("Beneficiary Status Report")
    doc.period_type = period_type
    doc.date_from = df
    doc.date_to = dt
    doc.title = f"Beneficiaries {period_type}: {df} → {dt}"
    doc.insert(ignore_permissions=True)
    doc.run_generation()
    # generate attachments for verification
    try:
        doc.generate_excel()
    except Exception:
        pass
    try:
        doc.generate_pdf()
    except Exception:
        pass
    return doc.name

