import json
from datetime import date, timedelta, datetime
from typing import Any, Dict, List, Tuple, Optional

import frappe
from frappe.model.document import Document
from frappe.utils import getdate, get_datetime, now_datetime

from assistant_crm.assistant_crm.doctype.sla_compliance_report.sla_compliance_report import aggregate_sla_compliance


class BranchPerformanceReport(Document):
    def before_insert(self):
        self._ensure_dates()

    def before_save(self):
        self._ensure_dates()

    def _ensure_dates(self):
        if not getattr(self, "date_from", None) or not getattr(self, "date_to", None):
            # Default: previous full month for Monthly, last 90 days for others
            today = getdate()
            if (self.period_type or "Monthly") == "Monthly":
                first_this_month = today.replace(day=1)
                last_prev_month = first_this_month - timedelta(days=1)
                first_prev_month = last_prev_month.replace(day=1)
                self.date_from = first_prev_month
                self.date_to = last_prev_month
            else:
                self.date_to = today
                self.date_from = today - timedelta(days=89)

    @frappe.whitelist()
    def run_generation(self) -> Dict[str, Any]:
        self._ensure_dates()
        filters = {
            "region": (self.region_filter or "All").strip() or "All",
            "branch": (self.branch_filter or "").strip() or None,
            "channel": None if (self.channel_filter or "All") == "All" else self.channel_filter,
            "priority": None if (self.priority_filter or "All") == "All" else self.priority_filter,
        }
        summary, branch_rows, region_rows, charts = aggregate_branch_performance(self.date_from, self.date_to, filters)

        # Copy summary metrics onto the document for UI, exports and AI
        self.total_branches = summary.get("total_branches", 0)
        self.total_issues = summary.get("total_issues", 0)
        self.total_claims = summary.get("total_claims", 0)
        self.total_complaints = summary.get("total_complaints", 0)
        self.total_escalations = summary.get("total_escalations", 0)
        self.overall_sla_compliance_percent = summary.get("overall_sla", 0.0)
        self.avg_issue_resolution_days = summary.get("avg_issue_resolution_days", 0.0)
        self.avg_claim_resolution_days = summary.get("avg_claim_resolution_days", 0.0)
        self.avg_complaint_resolution_days = summary.get("avg_complaint_resolution_days", 0.0)

        # Charts
        self.branch_overview_chart_json = json.dumps(charts.get("branch_overview") or {})
        self.sla_branch_chart_json = json.dumps(charts.get("sla_branch") or {})
        self.regional_comparison_chart_json = json.dumps(charts.get("regional") or {})

        trend_chart = build_branch_trend_chart(self)
        self.trend_chart_json = json.dumps(trend_chart or {})

        # Data blobs for UI/exports
        self.rows_json = json.dumps(branch_rows)
        self.filters_json = json.dumps(filters)
        self.report_html = build_report_html(self, summary, branch_rows, region_rows)
        self.generated_at = now_datetime()
        self.generated_by = frappe.session.user
        self.save()
        return {"ok": True, "summary": summary}

    @frappe.whitelist()
    def generate_pdf(self) -> str:
        from frappe.utils.pdf import get_pdf

        html = getattr(self, "report_html", None) or build_report_html(self, {}, [], [])
        filename = f"Branch_Performance_Report_{self.name}.pdf"
        file_doc = _attach_to_doc(self.doctype, self.name, filename, get_pdf(html))
        return getattr(file_doc, "file_url", None) or filename

    @frappe.whitelist()
    def generate_excel(self) -> str:
        rows = json.loads(self.rows_json or "[]")
        filename = f"Branch_Performance_Report_{self.name}.xlsx"
        content: bytes

        # Simple headers matching branch_rows structure
        headers = [
            "Branch",
            "Region",
            "Total Claims",
            "Total Claim Amount",
            "Claims Resolved",
            "Claims Rejected",
            "Total Complaints",
            "Complaints Escalated",
            "Complaints Resolved",
            "SLA %",
            "Avg Claim Resolution Days",
            "Avg Complaint Resolution Days",
        ]

        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            # Title + period
            ws["A1"] = "WCFCB Branch Performance Report"
            ws["A1"].font = Font(b=True, size=14)
            ws["A2"] = f"{self.period_type} | {self.date_from} to {self.date_to}"
            ws["A2"].font = Font(color="666666")

            # KPI table
            ws["A4"] = "Metric"; ws["B4"] = "Value"
            ws["A4"].font = ws["B4"].font = Font(b=True)
            metrics = [
                ("Total Branches", int(self.total_branches or 0)),
                ("Total Claims", int(self.total_claims or 0)),
                ("Total Complaints", int(self.total_complaints or 0)),
                ("Total Escalations", int(self.total_escalations or 0)),
                ("Overall SLA %", float(self.overall_sla_compliance_percent or 0.0)),
                ("Avg Claim Resolution Days", float(self.avg_claim_resolution_days or 0.0)),
                ("Avg Complaint Resolution Days", float(self.avg_complaint_resolution_days or 0.0)),
            ]
            r = 5
            for label, value in metrics:
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=value)
                r += 1

            # Colour legend for SLA tiers
            ws["D4"] = "SLA Legend"; ws["D4"].font = Font(b=True)
            legend = [
                ("High Performance (>= 90%)", "E9F7EF"),
                ("Watch List (75% - 89.9%)", "FEF9E7"),
                ("At Risk (< 75%)", "FDEDEC"),
            ]
            rr = 5
            for label, color in legend:
                ws.cell(row=rr, column=4, value=label)
                c = ws.cell(row=rr, column=5, value=" ")
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                rr += 1

            # Branch detail sheet
            ws_branch = wb.create_sheet(title="Branch Details")
            for ci, h in enumerate(headers, start=1):
                cell = ws_branch.cell(row=1, column=ci, value=h)
                cell.font = Font(b=True)

            for idx, row in enumerate(rows, start=2):
                vals = [
                    row.get("branch"),
                    row.get("region"),
                    int(row.get("total_claims", 0) or 0),
                    float(row.get("total_claim_amount", 0.0) or 0.0),
                    int(row.get("claims_resolved", 0) or 0),
                    int(row.get("claims_rejected", 0) or 0),
                    int(row.get("total_complaints", 0) or 0),
                    int(row.get("complaints_escalated", 0) or 0),
                    int(row.get("complaints_resolved", 0) or 0),
                    float(row.get("sla_percent", 0.0) or 0.0),
                    float(row.get("avg_claim_resolution_days", 0.0) or 0.0),
                    float(row.get("avg_complaint_resolution_days", 0.0) or 0.0),
                ]
                for ci, v in enumerate(vals, start=1):
                    ws_branch.cell(row=idx, column=ci, value=v)

                sla_val = float(row.get("sla_percent", 0.0) or 0.0)
                if sla_val >= 90.0:
                    fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")
                elif sla_val >= 75.0:
                    fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                else:
                    fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
                for ci in range(1, len(headers) + 1):
                    ws_branch.cell(row=idx, column=ci).fill = fill

            # Regional summary sheet (derived from branch rows)
            ws_region = wb.create_sheet(title="Regional Summary")
            region_headers = ["Region", "Total Claims", "Total Complaints", "SLA %"]
            for ci, h in enumerate(region_headers, start=1):
                cell = ws_region.cell(row=1, column=ci, value=h)
                cell.font = Font(b=True)

            region_rows = _build_region_aggregates(rows)
            r = 2
            for reg in region_rows:
                vals = [
                    reg.get("region"),
                    int(reg.get("total_claims", 0) or 0),
                    int(reg.get("total_complaints", 0) or 0),
                    float(reg.get("sla_percent", 0.0) or 0.0),
                ]
                for ci, v in enumerate(vals, start=1):
                    ws_region.cell(row=r, column=ci, value=v)
                r += 1

            # Trends sheet with historical KPIs
            ws_trend = wb.create_sheet(title="Trends")
            ws_trend["A1"] = "Period"; ws_trend["B1"] = "Total Claims"; ws_trend["C1"] = "Total Complaints"; ws_trend["D1"] = "Total Escalations"; ws_trend["E1"] = "SLA %"
            for ci in range(1, 6):
                ws_trend.cell(row=1, column=ci).font = Font(b=True)

            try:
                trend_points = _get_trend_points(self, max_points=12)
            except Exception:
                trend_points = []

            r = 2
            for p in trend_points:
                vals = [
                    p.get("label"),
                    int(p.get("total_claims", 0) or 0),
                    int(p.get("total_complaints", 0) or 0),
                    int(p.get("total_escalations", 0) or 0),
                    float(p.get("sla_percent", 0.0) or 0.0),
                ]
                for ci, v in enumerate(vals, start=1):
                    ws_trend.cell(row=r, column=ci, value=v)
                r += 1

            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()
        except Exception:
            # Fallback to a simple single-sheet export via xlsxutils/CSV
            simple_rows = [
                [
                    r.get("branch"),
                    r.get("region"),
                    r.get("total_claims"),
                    r.get("total_complaints"),
                    r.get("sla_percent"),
                ]
                for r in rows
            ]
            try:
                from frappe.utils.xlsxutils import make_xlsx  # type: ignore

                xlsx_file = make_xlsx([["Branch", "Region", "Total Claims", "Total Complaints", "SLA %"]] + simple_rows, "Branch Performance")
                content = xlsx_file.getvalue()
            except Exception:
                import io as _io_mod, csv

                buf = _io_mod.StringIO()
                w = csv.writer(buf)
                w.writerow(["Branch", "Region", "Total Claims", "Total Complaints", "SLA %"])
                w.writerows(simple_rows)
                content = buf.getvalue().encode()
                filename = filename.replace(".xlsx", ".csv")

        file_doc = _attach_to_doc(self.doctype, self.name, filename, content)
        return getattr(file_doc, "file_url", None) or filename

    @frappe.whitelist()
    def email_report(self, recipients: Optional[List[str]] = None) -> Dict[str, Any]:
        """Email the Branch Performance Report to Senior Management & Branch Managers.

        If recipients are not provided, all enabled users with either the
        "Senior Management" or "Branch Manager" role will be used.
        """
        recips = recipients or _get_role_emails(["Senior Management", "Branch Manager"]) or []
        if not recips:
            return {"message": "no-recipients"}

        try:
            # Ensure latest aggregation and attachments
            self.run_generation()
            try:
                self.generate_pdf()
            except Exception:
                frappe.log_error(frappe.get_traceback(), "Branch Performance Report: generate_pdf in email_report")
            try:
                self.generate_excel()
            except Exception:
                frappe.log_error(frappe.get_traceback(), "Branch Performance Report: generate_excel in email_report")

            files = frappe.get_all(
                "File",
                filters={
                    "attached_to_doctype": self.doctype,
                    "attached_to_name": self.name,
                },
                fields=["name"],
                limit=20,
            )
            attachments = []
            for f in files:
                try:
                    file_doc = frappe.get_doc("File", f.get("name"))
                    attachments.append({
                        "fname": file_doc.file_name,
                        "fcontent": file_doc.get_content(),
                    })
                except Exception:
                    frappe.log_error(frappe.get_traceback(), "Branch Performance Report: load attachment in email_report")

            frappe.sendmail(
                recipients=recips,
                subject=f"WCFCB Branch Performance Report: {self.period_type} ({self.date_from} â†’ {self.date_to})",
                message="Attached is the latest Branch Performance Report.",
                attachments=attachments,
            )
            return {"message": "sent", "recipients": recips}
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Branch Performance Report: email_report")
            frappe.throw(f"Failed to email Branch Performance Report: {e}")


@frappe.whitelist()
def get_ai_insights(name: str, query: str) -> Dict[str, Any]:
    """Return WorkCom-style insights for a Branch Performance Report.

    Builds a compact JSON context with the current window KPIs, per-branch rows,
    regional breakdowns and recent history, and passes it to WorkCom via EnhancedAIService.
    """
    doc = frappe.get_doc("Branch Performance Report", name)

    # Fetch recent history for trends
    history = frappe.get_all(
        "Branch Performance Report",
        filters={"period_type": doc.period_type},
        fields=[
            "name",
            "date_from",
            "date_to",
            "total_branches",
            "total_claims",
            "total_complaints",
            "total_escalations",
            "overall_sla_compliance_percent",
            "avg_claim_resolution_days",
            "avg_complaint_resolution_days",
        ],
        order_by="date_from desc",
        limit=12,
    )

    context = {
        "window": {
            "period_type": doc.period_type,
            "from": str(doc.date_from),
            "to": str(doc.date_to),
        },
        "current": {
            "total_branches": doc.total_branches,
            "total_claims": doc.total_claims,
            "total_complaints": doc.total_complaints,
            "total_escalations": doc.total_escalations,
            "overall_sla_percent": doc.overall_sla_compliance_percent,
            "avg_claim_resolution_days": doc.avg_claim_resolution_days,
            "avg_complaint_resolution_days": doc.avg_complaint_resolution_days,
            "rows": json.loads(doc.rows_json or "[]"),
        },
        "history": history,
    }

    try:
        from assistant_crm.services.enhanced_ai_service import EnhancedAIService

        ai = EnhancedAIService()
        answer = ai.generate_branch_performance_report_insights(query=query, context=context)
        return {"insights": answer}
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Branch Performance Report AI Insights Error")
        return {
            "insights": (
                "AI insights are temporarily unavailable. Please ask your system "
                "administrator to configure WorkCom/OpenAI settings in Enhanced AI Settings."
            )
        }




@frappe.whitelist()
def generate_pdf_file(name: str) -> Dict[str, str]:
    """Generate and attach a PDF for the given Branch Performance Report and return its URL."""
    doc = frappe.get_doc("Branch Performance Report", name)
    file_url = doc.generate_pdf()
    return {"file_url": file_url}


@frappe.whitelist()
def generate_excel_file(name: str) -> Dict[str, str]:
    """Generate and attach an Excel file for the given Branch Performance Report and return its URL."""
    doc = frappe.get_doc("Branch Performance Report", name)
    file_url = doc.generate_excel()
    return {"file_url": file_url}

def aggregate_branch_performance(date_from: Any, date_to: Any, filters: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    df = getdate(date_from)
    dt = getdate(date_to)

    branch_filter = (filters.get("branch") or "").strip().lower()
    region_filter = (filters.get("region") or "All").strip() or "All"
    channel = filters.get("channel")
    priority = filters.get("priority")

    claims_by_branch = _aggregate_claims_by_branch(df, dt, branch_filter, priority)
    complaints_by_branch = _aggregate_complaints_by_branch(df, dt, branch_filter, channel, priority)
    issues_by_branch = _aggregate_issues_by_branch(df, dt, branch_filter, priority)

    # Reuse SLA aggregation from SLA Compliance Report for within/breached counts
    sla_counts, sla_charts, _ = aggregate_sla_compliance(df, dt, {
        "branch": filters.get("branch"),
        "role": None,
        "channel": channel,
        "priority": priority,
    })
    sla_by_branch = _extract_sla_branch_stats(sla_charts)

    all_branches = sorted(set(claims_by_branch.keys()) | set(complaints_by_branch.keys()) | set(sla_by_branch.keys()) | set(issues_by_branch.keys()))

    branch_rows: List[Dict[str, Any]] = []
    region_acc: Dict[str, Dict[str, Any]] = {}

    for branch in all_branches:
        region = _map_branch_to_region(branch)
        if region_filter not in ("All", "", None) and region != region_filter:
            continue

        c = claims_by_branch.get(branch, {})
        comp = complaints_by_branch.get(branch, {})
        iss = issues_by_branch.get(branch, {})
        sla = sla_by_branch.get(branch, {"within": 0, "breached": 0})
        within = int(sla.get("within", 0))
        breached = int(sla.get("breached", 0))
        total_sla = within + breached
        sla_percent = (within / total_sla * 100.0) if total_sla else 0.0

        avg_claim_days = _avg_from_sums(c.get("resolution_days_sum", 0.0), c.get("resolution_days_count", 0))
        avg_comp_days = _avg_from_sums(comp.get("resolution_days_sum", 0.0), comp.get("resolution_days_count", 0))
        avg_issue_days = _avg_from_sums(iss.get("resolution_days_sum", 0.0), iss.get("resolution_days_count", 0))

        row = {
            "branch": branch,
            "region": region,
            "total_claims": int(c.get("total", 0)),
            "total_claim_amount": float(c.get("amount_total", 0.0)),
            "claims_resolved": int(c.get("resolved", 0)),
            "claims_rejected": int(c.get("rejected", 0)),
            "total_complaints": int(comp.get("total", 0)),
            "complaints_escalated": int(comp.get("escalated", 0)),
            "complaints_resolved": int(comp.get("resolved", 0)),
            "total_issues": int(iss.get("total", 0)),
            "issues_open": int(iss.get("open", 0)),
            "issues_resolved": int(iss.get("resolved", 0)),
            "sla_within": within,
            "sla_breached": breached,
            "sla_percent": sla_percent,
            "avg_claim_resolution_days": avg_claim_days,
            "avg_complaint_resolution_days": avg_comp_days,
            "avg_issue_resolution_days": avg_issue_days,
        }
        branch_rows.append(row)

        reg_key = region or "Unassigned"
        reg = region_acc.setdefault(reg_key, {
            "region": reg_key,
            "total_claims": 0,
            "total_claim_amount": 0.0,
            "total_complaints": 0,
            "total_issues": 0,
            "sla_within": 0,
            "sla_breached": 0,
        })
        reg["total_claims"] += row["total_claims"]
        reg["total_claim_amount"] += row["total_claim_amount"]
        reg["total_complaints"] += row["total_complaints"]
        reg["total_issues"] += row["total_issues"]
        reg["sla_within"] += within
        reg["sla_breached"] += breached

    branch_rows.sort(key=lambda x: ((x.get("region") or ""), (x.get("branch") or "")))

    region_rows: List[Dict[str, Any]] = []
    for reg_key, data in sorted(region_acc.items(), key=lambda kv: kv[0]):
        total_sla = data["sla_within"] + data["sla_breached"]
        data["sla_percent"] = (data["sla_within"] / total_sla * 100.0) if total_sla else 0.0
        region_rows.append(data)

    total_branches = len(branch_rows)
    total_claims = sum(r["total_claims"] for r in branch_rows)
    total_complaints = sum(r["total_complaints"] for r in branch_rows)
    total_issues = sum(r["total_issues"] for r in branch_rows)
    total_escalations = sum(r["complaints_escalated"] for r in branch_rows)
    avg_claim_days = _avg([r["avg_claim_resolution_days"] for r in branch_rows])
    avg_comp_days = _avg([r["avg_complaint_resolution_days"] for r in branch_rows])
    avg_issue_days = _avg([r["avg_issue_resolution_days"] for r in branch_rows])

    summary: Dict[str, Any] = {
        "total_branches": total_branches,
        "total_claims": total_claims,
        "total_complaints": total_complaints,
        "total_issues": total_issues,
        "total_escalations": total_escalations,
        "overall_sla": float(sla_counts.get("compliance_percent", 0.0)),
        "avg_claim_resolution_days": avg_claim_days,
        "avg_complaint_resolution_days": avg_comp_days,
        "avg_issue_resolution_days": avg_issue_days,
    }

    charts: Dict[str, Any] = {
        "branch_overview": _build_branch_overview_chart(branch_rows),
        "sla_branch": _build_sla_branch_chart(branch_rows),
        "regional": _build_regional_chart(region_rows),
    }

    return summary, branch_rows, region_rows, charts



def _aggregate_claims_by_branch(df: date, dt: date, branch_filter: str, priority: Optional[str]) -> Dict[str, Dict[str, Any]]:
    # Ensure dates are properly formatted as strings for SQL query
    from_date = str(df) if df else None
    to_date = str(dt) if dt else None

    # Use raw SQL to avoid Frappe's filter conversion issues with nullable date fields
    # Now includes branch field directly from Claim
    # Include all claims (docstatus >= 0) not just submitted ones
    sql = """
        SELECT name, branch, approved_by, status, submitted_date, approved_on, amount, creation
        FROM `tabClaim`
        WHERE docstatus >= 0
    """
    params = []

    if from_date and to_date:
        # Use creation date if submitted_date is null
        sql += " AND (submitted_date BETWEEN %s AND %s OR (submitted_date IS NULL AND creation BETWEEN %s AND %s))"
        params.extend([from_date, to_date, from_date, to_date])

    if priority:
        sql += " AND priority = %s"
        params.append(priority)

    # Filter by branch if specified
    if branch_filter:
        sql += " AND branch LIKE %s"
        params.append(f"%{branch_filter}%")

    sql += " LIMIT 10000"

    rows = frappe.db.sql(sql, params, as_dict=True)

    result: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        # Use branch field directly, fall back to deriving from approved_by if not set
        branch = r.get("branch") or _derive_branch_for_user(r.get("approved_by"))
        if not branch:
            branch = "Unassigned"

        bucket = result.setdefault(
            branch,
            {
                "total": 0,
                "resolved": 0,
                "rejected": 0,
                "amount_total": 0.0,
                "resolution_days_sum": 0.0,
                "resolution_days_count": 0,
            },
        )
        bucket["total"] += 1
        bucket["amount_total"] += float(r.get("amount") or 0.0)

        status = (r.get("status") or "").lower()
        if status in {"approved", "paid", "settled", "closed"}:
            bucket["resolved"] += 1
        if status in {"rejected", "declined"}:
            bucket["rejected"] += 1

        submitted = r.get("submitted_date")
        approved_on = r.get("approved_on")
        if submitted and approved_on:
            try:
                start = get_datetime(submitted)
                end = get_datetime(approved_on)
                days = (end - start).days
                bucket["resolution_days_sum"] += max(days, 0)
                bucket["resolution_days_count"] += 1
            except Exception:
                pass

    return result


def _aggregate_complaints_by_branch(df: date, dt: date, branch_filter: str, channel: Optional[str], priority: Optional[str]) -> Dict[str, Dict[str, Any]]:
    # Ensure dates are properly formatted as strings for SQL query
    from_date = str(df) if df else None
    to_date = str(dt) if dt else None

    # Use raw SQL to avoid Frappe's filter conversion issues
    # Now includes branch field directly from Unified Inbox Conversation
    sql = """
        SELECT name, branch, assigned_agent, status, priority, creation_time, last_message_time, escalated_at
        FROM `tabUnified Inbox Conversation`
        WHERE 1=1
    """
    params = []

    if from_date and to_date:
        sql += " AND creation >= %s AND creation <= %s"
        params.extend([from_date, to_date])

    if channel:
        sql += " AND platform = %s"
        params.append(channel)
    if priority:
        sql += " AND priority = %s"
        params.append(priority)

    # Filter by branch if specified (case-insensitive partial match)
    if branch_filter:
        sql += " AND LOWER(COALESCE(branch, '')) LIKE %s"
        params.append(f"%{branch_filter.lower()}%")

    sql += " LIMIT 10000"

    convs = frappe.db.sql(sql, params, as_dict=True)

    result: Dict[str, Dict[str, Any]] = {}
    for c in convs:
        # Use branch field directly, fall back to deriving from assigned_agent if not set
        branch = c.get("branch") or _derive_branch_for_user(c.get("assigned_agent"))
        if not branch:
            branch = "Unassigned"

        bucket = result.setdefault(
            branch,
            {
                "total": 0,
                "resolved": 0,
                "escalated": 0,
                "resolution_days_sum": 0.0,
                "resolution_days_count": 0,
            },
        )
        bucket["total"] += 1
        status = (c.get("status") or "").lower()
        if status in {"closed", "resolved"}:
            bucket["resolved"] += 1
        if c.get("escalated_at"):
            bucket["escalated"] += 1

        # Approximate resolution time as time between creation and last message
        creation_time = c.get("creation_time")
        last_msg_time = c.get("last_message_time")
        if creation_time and last_msg_time and status in {"closed", "resolved"}:
            try:
                start = get_datetime(creation_time)
                end = get_datetime(last_msg_time)
                days = (end - start).total_seconds() / (60.0 * 60.0 * 24.0)
                bucket["resolution_days_sum"] += max(days, 0.0)
                bucket["resolution_days_count"] += 1
            except Exception:
                pass

    return result


def _aggregate_issues_by_branch(df: date, dt: date, branch_filter: str, priority: Optional[str]) -> Dict[str, Dict[str, Any]]:
    """Aggregate Issue tickets by custom_branch field"""
    from_date = str(df) if df else None
    to_date = str(dt) if dt else None

    sql = """
        SELECT name, custom_branch, status, priority, creation, resolution_date
        FROM `tabIssue`
        WHERE 1=1
    """
    params = []

    if from_date and to_date:
        sql += " AND creation >= %s AND creation <= %s"
        params.extend([from_date, to_date])

    if priority:
        sql += " AND priority = %s"
        params.append(priority)

    # Filter by branch if specified (case-insensitive partial match)
    if branch_filter:
        sql += " AND LOWER(COALESCE(custom_branch, '')) LIKE %s"
        params.append(f"%{branch_filter.lower()}%")

    sql += " LIMIT 10000"

    issues = frappe.db.sql(sql, params, as_dict=True)

    result: Dict[str, Dict[str, Any]] = {}
    for i in issues:
        branch = i.get("custom_branch") or "Unassigned"

        bucket = result.setdefault(
            branch,
            {
                "total": 0,
                "open": 0,
                "resolved": 0,
                "resolution_days_sum": 0.0,
                "resolution_days_count": 0,
            },
        )
        bucket["total"] += 1
        status = (i.get("status") or "").lower()
        if status in {"closed", "resolved"}:
            bucket["resolved"] += 1
        elif status in {"open", "replied"}:
            bucket["open"] += 1

        # Calculate resolution time
        creation = i.get("creation")
        resolution_date = i.get("resolution_date")
        if creation and resolution_date and status in {"closed", "resolved"}:
            try:
                start = get_datetime(creation)
                end = get_datetime(resolution_date)
                days = (end - start).days
                bucket["resolution_days_sum"] += max(days, 0)
                bucket["resolution_days_count"] += 1
            except Exception:
                pass

    return result


def _derive_branch_for_user(user_id: Optional[str]) -> str:
    if not user_id:
        return "Unassigned"
    try:
        user = frappe.get_cached_doc("User", user_id)
        meta = user.meta
        if meta.has_field("branch") and getattr(user, "branch", None):
            return str(user.branch)
        if meta.has_field("department") and getattr(user, "department", None):
            return str(user.department)
    except Exception:
        return "Unassigned"
    return "Unassigned"


_REGION_NAMES = [
    "Lusaka",
    "Copperbelt",
    "Northern",
    "Eastern",
    "Southern",
    "Western",
    "Central",
    "Luapula",
    "Muchinga",
    "North-Western",
]


def _map_branch_to_region(branch: Optional[str]) -> str:
    if not branch:
        return "Unassigned"
    b = branch.lower()
    for region in _REGION_NAMES:
        if region.lower() in b:
            return region
    return "Other"


def _avg_from_sums(total: float, count: int) -> float:
    return float(total) / count if count else 0.0


def _avg(values: List[float]) -> float:
    vals = [v for v in values if v]
    return float(sum(vals)) / len(vals) if vals else 0.0


def _extract_sla_branch_stats(charts: Dict[str, Any]) -> Dict[str, Dict[str, int]]:
    """Pull branch-level within/breached counts from SLA charts structure."""
    result: Dict[str, Dict[str, int]] = {}
    branch_chart = (charts or {}).get("branch") or {}
    data = branch_chart.get("data") or {}
    labels = data.get("labels") or []
    datasets = data.get("datasets") or []
    within_ds = next((d for d in datasets if (d.get("name") or "").lower().startswith("within")), None)
    breached_ds = next((d for d in datasets if (d.get("name") or "").lower().startswith("breach")), None)
    within_vals = within_ds.get("values") if isinstance(within_ds, dict) else []
    breached_vals = breached_ds.get("values") if isinstance(breached_ds, dict) else []
    for idx, label in enumerate(labels):
        result[label] = {
            "within": int((within_vals[idx] if idx < len(within_vals) else 0) or 0),
            "breached": int((breached_vals[idx] if idx < len(breached_vals) else 0) or 0),
        }
    return result


def _build_branch_overview_chart(branch_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not branch_rows:
        return {}
    top = sorted(
        branch_rows,
        key=lambda r: (r.get("total_claims", 0) or 0) + (r.get("total_complaints", 0) or 0) + (r.get("total_issues", 0) or 0),
        reverse=True,
    )[:12]
    labels = [r.get("branch") or "" for r in top]
    return {
        "type": "bar",
        "data": {
            "labels": labels,
            "datasets": [
                {"name": "Issues", "values": [int(r.get("total_issues", 0) or 0) for r in top]},
                {"name": "Claims", "values": [int(r.get("total_claims", 0) or 0) for r in top]},
                {"name": "Complaints", "values": [int(r.get("total_complaints", 0) or 0) for r in top]},
            ],
        },
        "barOptions": {"stacked": 1},
    }


def _build_sla_branch_chart(branch_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not branch_rows:
        return {}
    top = sorted(branch_rows, key=lambda r: float(r.get("sla_percent", 0.0) or 0.0), reverse=True)[:12]
    labels = [r.get("branch") or "" for r in top]
    values = [round(float(r.get("sla_percent", 0.0) or 0.0), 2) for r in top]
    return {
        "type": "bar",
        "data": {
            "labels": labels,
            "datasets": [{"name": "SLA %", "values": values}],
        },
    }


def _build_regional_chart(region_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not region_rows:
        return {}
    rows = region_rows
    labels = [r.get("region") or "" for r in rows]
    return {
        "type": "bar",
        "data": {
            "labels": labels,
            "datasets": [
                {"name": "Claims", "values": [int(r.get("total_claims", 0) or 0) for r in rows]},
                {"name": "Complaints", "values": [int(r.get("total_complaints", 0) or 0) for r in rows]},
                {"name": "SLA %", "values": [round(float(r.get("sla_percent", 0.0) or 0.0), 2) for r in rows]},
            ],
        },
    }


def _build_region_aggregates(branch_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    acc: Dict[str, Dict[str, Any]] = {}
    for row in branch_rows:
        region = row.get("region") or "Unassigned"
        reg = acc.setdefault(
            region,
            {"region": region, "total_claims": 0, "total_complaints": 0, "sla_within": 0, "sla_breached": 0},
        )
        reg["total_claims"] += int(row.get("total_claims", 0) or 0)
        reg["total_complaints"] += int(row.get("total_complaints", 0) or 0)
        reg["sla_within"] += int(row.get("sla_within", 0) or 0)
        reg["sla_breached"] += int(row.get("sla_breached", 0) or 0)
    region_rows: List[Dict[str, Any]] = []
    for reg_key, data in sorted(acc.items(), key=lambda kv: kv[0]):
        total_sla = data["sla_within"] + data["sla_breached"]
        data["sla_percent"] = (data["sla_within"] / total_sla * 100.0) if total_sla else 0.0
        region_rows.append(data)
    return region_rows


def _get_trend_points(doc: BranchPerformanceReport, max_points: int = 6) -> List[Dict[str, Any]]:
    """Return historical KPI points including the current report window."""
    base = getdate(doc.date_from) if doc.date_from else getdate()
    period_type = doc.period_type or "Monthly"

    points: List[Dict[str, Any]] = []
    current_start = base
    for _ in range(max_points):
        if period_type == "Quarterly":
            # Go back in 3-month steps
            month = ((current_start.month - 1) // 3) * 3 + 1
            start = current_start.replace(month=month, day=1)
            end_month = month + 2
            end_year = current_start.year
            if end_month > 12:
                end_month -= 12
                end_year += 1
            end = date(end_year, end_month, 1)
            end = (end.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            label = f"Q{((month - 1) // 3) + 1} {start.year}"
        else:
            start = current_start.replace(day=1)
            end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            label = start.strftime("%b %Y")

        prev = frappe.get_all(
            "Branch Performance Report",
            filters={"period_type": period_type, "date_from": ("<=", start), "date_to": (">=", end)},
            fields=[
                "name",
                "total_claims",
                "total_complaints",
                "total_escalations",
                "overall_sla_compliance_percent",
            ],
            order_by="creation desc",
            limit=1,
        )
        if prev:
            row = prev[0]
            points.append(
                {
                    "label": label,
                    "total_claims": row.get("total_claims") or 0,
                    "total_complaints": row.get("total_complaints") or 0,
                    "total_escalations": row.get("total_escalations") or 0,
                    "sla_percent": float(row.get("overall_sla_compliance_percent") or 0.0),
                }
            )
        current_start = start - timedelta(days=1)

    if not points:
        points.append(
            {
                "label": f"{period_type} {base}",
                "total_claims": int(doc.total_claims or 0),
                "total_complaints": int(doc.total_complaints or 0),
                "total_escalations": int(doc.total_escalations or 0),
                "sla_percent": float(doc.overall_sla_compliance_percent or 0.0),
            }
        )

    points.reverse()
    return points


def build_branch_trend_chart(doc: BranchPerformanceReport) -> Dict[str, Any]:
    points = _get_trend_points(doc, max_points=6)
    labels = [p["label"] for p in points]
    return {
        "type": "line",
        "data": {
            "labels": labels,
            "datasets": [
                {"name": "Claims", "values": [p["total_claims"] for p in points]},
                {"name": "Complaints", "values": [p["total_complaints"] for p in points]},
                {"name": "Escalations", "values": [p["total_escalations"] for p in points]},
                {"name": "SLA %", "values": [p["sla_percent"] for p in points]},
            ],
        },
    }



def build_report_html(doc: BranchPerformanceReport, summary: Dict[str, Any], branches: List[Dict[str, Any]], regions: List[Dict[str, Any]]) -> str:
    """Build interactive HTML report with clickable KPIs and branch detail table"""
    from urllib.parse import urlencode

    date_from = str(getattr(doc, "date_from", "")) or ""
    date_to = str(getattr(doc, "date_to", "")) or ""
    branch_filter = (getattr(doc, "branch_filter", "") or "").strip()

    total_branches = summary.get("total_branches", getattr(doc, "total_branches", 0))
    total_claims = summary.get("total_claims", getattr(doc, "total_claims", 0))
    total_complaints = summary.get("total_complaints", getattr(doc, "total_complaints", 0))
    total_issues = summary.get("total_issues", 0)
    total_escalations = summary.get("total_escalations", getattr(doc, "total_escalations", 0))
    sla = summary.get("overall_sla", getattr(doc, "overall_sla_compliance_percent", 0))
    avg_claim_days = summary.get("avg_claim_resolution_days", getattr(doc, "avg_claim_resolution_days", 0))
    avg_complaint_days = summary.get("avg_complaint_resolution_days", getattr(doc, "avg_complaint_resolution_days", 0))
    avg_issue_days = summary.get("avg_issue_resolution_days", 0)

    # Build filter URLs for clickable KPIs
    from urllib.parse import quote
    encoded_branch = quote(branch_filter) if branch_filter else ""

    claims_url = f"/app/claim?branch={encoded_branch}" if branch_filter else "/app/claim"
    conversations_url = f"/app/unified-inbox-conversation?branch={encoded_branch}" if branch_filter else "/app/unified-inbox-conversation"
    issues_url = f"/app/issue?custom_branch={encoded_branch}" if branch_filter else "/app/issue"

    # KPI cards with clickable links
    html = f"""
    <style>
        .br-report-container {{ font-family: inherit; }}
        .br-kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin-bottom: 24px; }}
        .br-kpi-card {{ background: var(--card-bg); border: 1px solid var(--border-color); border-radius: 8px; padding: 16px; text-align: center; }}
        .br-kpi-card a {{ text-decoration: none; color: inherit; }}
        .br-kpi-card:hover {{ box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        .br-kpi-value {{ font-size: 28px; font-weight: 600; color: var(--primary); }}
        .br-kpi-value a {{ color: var(--primary); }}
        .br-kpi-value a:hover {{ text-decoration: underline; }}
        .br-kpi-label {{ font-size: 12px; color: var(--text-muted); margin-top: 4px; }}
        .br-table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
        .br-table th, .br-table td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid var(--border-color); }}
        .br-table th {{ background: var(--bg-color); font-weight: 600; font-size: 12px; text-transform: uppercase; }}
        .br-table tr:hover {{ background: var(--bg-light-gray); }}
        .br-table td a {{ color: var(--primary); text-decoration: none; }}
        .br-table td a:hover {{ text-decoration: underline; }}
        .br-section-title {{ font-size: 16px; font-weight: 600; margin: 24px 0 12px 0; color: var(--heading-color); }}
        .br-sla-good {{ color: var(--green-500); }}
        .br-sla-warning {{ color: var(--yellow-500); }}
        .br-sla-bad {{ color: var(--red-500); }}
    </style>
    <div class="br-report-container">
        <h3>Branch Performance Summary</h3>
        <p style="color: var(--text-muted); font-size: 13px;">
            Period: {date_from} to {date_to}
            {f' | Branch: {branch_filter}' if branch_filter else ''}
        </p>

        <div class="br-kpi-grid">
            <div class="br-kpi-card">
                <div class="br-kpi-value">{total_branches}</div>
                <div class="br-kpi-label">Branches</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value"><a href="{issues_url}" target="_blank">{total_issues}</a></div>
                <div class="br-kpi-label">Total Issues</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value"><a href="{claims_url}" target="_blank">{total_claims}</a></div>
                <div class="br-kpi-label">Total Claims</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value"><a href="{conversations_url}" target="_blank">{total_complaints}</a></div>
                <div class="br-kpi-label">Total Complaints</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value">{total_escalations}</div>
                <div class="br-kpi-label">Escalations</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value">{sla:.1f}%</div>
                <div class="br-kpi-label">Overall SLA</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value">{avg_issue_days:.1f}</div>
                <div class="br-kpi-label">Avg Issue Days</div>
            </div>
            <div class="br-kpi-card">
                <div class="br-kpi-value">{avg_claim_days:.1f}</div>
                <div class="br-kpi-label">Avg Claim Days</div>
            </div>
        </div>
    """

    # Branch detail table with drill-down links
    if branches:
        html += """
        <div class="br-section-title">Branch Details</div>
        <table class="br-table">
            <thead>
                <tr>
                    <th>Branch</th>
                    <th>Region</th>
                    <th>Issues</th>
                    <th>Claims</th>
                    <th>Complaints</th>
                    <th>SLA %</th>
                    <th>Avg Issue (days)</th>
                </tr>
            </thead>
            <tbody>
        """
        for b in branches:
            branch_name = b.get("branch", "Unknown")
            region_name = b.get("region", "")
            issues = b.get("total_issues", 0)
            claims = b.get("total_claims", 0)
            complaints = b.get("total_complaints", 0)
            sla_pct = b.get("sla_percent", 0)
            avg_issue_res = b.get("avg_issue_resolution_days", 0)

            # SLA color coding
            if sla_pct >= 90:
                sla_class = "br-sla-good"
            elif sla_pct >= 75:
                sla_class = "br-sla-warning"
            else:
                sla_class = "br-sla-bad"

            # Build drill-down URLs for this specific branch (URL encode branch name)
            from urllib.parse import quote
            encoded_branch = quote(branch_name)
            branch_claims_url = f"/app/claim?branch={encoded_branch}"
            branch_issues_url = f"/app/issue?custom_branch={encoded_branch}"

            html += f"""
                <tr>
                    <td><strong><a href="{branch_issues_url}" target="_blank">{branch_name}</a></strong></td>
                    <td>{region_name}</td>
                    <td><a href="{branch_issues_url}" target="_blank">{issues}</a></td>
                    <td><a href="{branch_claims_url}" target="_blank">{claims}</a></td>
                    <td>{complaints}</td>
                    <td class="{sla_class}">{sla_pct:.1f}%</td>
                    <td>{avg_issue_res:.1f}</td>
                </tr>
            """
        html += """
            </tbody>
        </table>
        """

    html += "</div>"
    return html


def _attach_to_doc(doctype: str, name: str, filename: str, content: bytes):
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "attached_to_doctype": doctype,
        "attached_to_name": name,
        "content": content,
        "is_private": 1,
    })
    file_doc.save(ignore_permissions=True)
    return file_doc



def _get_role_emails(roles: List[str]) -> List[str]:
    if not roles:
        return []
    has_roles = frappe.get_all(
        "Has Role",
        filters={"role": ["in", roles]},
        fields=["parent"],
        limit=1000,
    )
    user_ids = {r["parent"] for r in has_roles if r.get("parent")}
    if not user_ids:
        return []
    users = frappe.get_all(
        "User",
        filters={"name": ["in", list(user_ids)], "enabled": 1, "user_type": "System User"},
        fields=["email"],
        limit=1000,
    )
    return sorted({u.get("email") for u in users if u.get("email")})


def _is_first_business_day(d: date) -> bool:
    """Return True if the given date is the first business day (Mon-Fri) of its month."""
    first = frappe.utils.get_first_day(d)
    first_bd = first
    while first_bd.weekday() > 4:  # 0=Mon, 6=Sun
        first_bd = frappe.utils.add_days(first_bd, 1)
    return d == first_bd


def _previous_month(d: date) -> Tuple[int, int]:
    """Return (month, year) for the calendar month immediately before the given date."""
    prev = (d.replace(day=1) - timedelta(days=1))
    return prev.month, prev.year


def _previous_quarter_range(d: date) -> Tuple[date, date]:
    """Return (start_date, end_date) for the quarter immediately before the given date."""
    q_month = ((d.month - 1) // 3) * 3 + 1
    current_q_start = date(d.year, q_month, 1)
    prev_q_end = current_q_start - timedelta(days=1)
    prev_q_month = ((prev_q_end.month - 1) // 3) * 3 + 1
    prev_q_start = date(prev_q_end.year, prev_q_month, 1)
    return prev_q_start, prev_q_end


@frappe.whitelist()
def schedule_monthly_branch_performance_reports():
    """Scheduler entry: on the first business day, generate previous month and email.

    Wired from hooks using a cron pattern that runs every weekday at 07:55.
    """
    today = getdate()
    if not _is_first_business_day(today):
        return

    mon, yr = _previous_month(today)
    df = date(yr, mon, 1)
    dt = frappe.utils.get_last_day(df)

    doc = frappe.new_doc("Branch Performance Report")
    doc.period_type = "Monthly"
    doc.date_from = df
    doc.date_to = dt
    doc.title = f"Branch Performance Monthly: {df} to {dt}"
    doc.insert(ignore_permissions=True)
    doc.run_generation()

    try:
        doc.email_report()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Branch Performance Report monthly email failed")


@frappe.whitelist()
def schedule_quarterly_branch_performance_reports():
    """Scheduler entry: on the first day after quarter end (Jan/Apr/Jul/Oct 1st) at 08:05.

    Generates a report for the previous full quarter and emails it to Senior Management
    and Branch Managers.
    """
    today = getdate()
    q_df, q_dt = _previous_quarter_range(today)

    doc = frappe.new_doc("Branch Performance Report")
    doc.period_type = "Quarterly"
    doc.date_from = q_df
    doc.date_to = q_dt
    doc.title = f"Branch Performance Quarterly: {q_df} to {q_dt}"
    doc.insert(ignore_permissions=True)
    doc.run_generation()

    try:
        doc.email_report()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Branch Performance Report quarterly email failed")


@frappe.whitelist()
def smoke_create_and_generate_branch_performance_report(period_type: str = "Monthly") -> Dict[str, Any]:
    """Create a test Branch Performance Report, run aggregation and exports, and return a summary.

    Intended to be invoked via `bench execute` for smoke testing. For "Monthly" it
    targets the previous full month; for anything else it targets the previous full
    quarter.
    """
    today = date.today()
    if (period_type or "").lower().startswith("month"):
        first_this_month = date(today.year, today.month, 1)
        last_prev_month = first_this_month - timedelta(days=1)
        df = date(last_prev_month.year, last_prev_month.month, 1)
        dt = last_prev_month
        ptype = "Monthly"
    else:
        df, dt = _previous_quarter_range(today)
        ptype = "Quarterly"

    doc = frappe.get_doc({
        "doctype": "Branch Performance Report",
        "period_type": ptype,
        "date_from": df,
        "date_to": dt,
        "title": f"Smoke {ptype} Branch Performance {df} to {dt}",
    })
    doc.insert(ignore_permissions=True)
    doc.run_generation()

    # Best-effort exports; failures should not break the smoke test
    try:
        doc.generate_excel()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Branch Performance Report smoke: generate_excel failed")
    try:
        doc.generate_pdf()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Branch Performance Report smoke: generate_pdf failed")

    return {
        "name": doc.name,
        "period_type": doc.period_type,
        "date_from": str(doc.date_from),
        "date_to": str(doc.date_to),
        "total_branches": getattr(doc, "total_branches", None),
        "total_claims": getattr(doc, "total_claims", None),
        "total_complaints": getattr(doc, "total_complaints", None),
        "total_escalations": getattr(doc, "total_escalations", None),
        "overall_sla_percent": getattr(doc, "overall_sla_compliance_percent", None),
    }


